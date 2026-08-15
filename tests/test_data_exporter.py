# Copyright 2026 Vsevolod Belonogov
# SPDX-License-Identifier: Apache-2.0

import csv
import sys
from pathlib import Path

import pytest

SRC = Path(__file__).resolve().parents[1] / "src"
if str(SRC) not in sys.path:
    sys.path.insert(0, str(SRC))

import data_exporter
from data_exporter import SQLiteDataExporter
from db_controller import SQLiteDatabaseController
from db_schema import DatabaseFileRegistry, DatabasePageLog
from schemas import FileSummary, PageResult, Status

import openpyxl



FILES = [
    (1, "docs/report.pdf", ".pdf", 3, Status.OK.value, "done", "1-3", "ok"),
    (2, "pics/photo.png", ".png", 1, Status.FAILURE.value, "corrupt header", "", "failure"),
    (3, "сканы/отчёт №3, «финал».pdf", ".pdf", 1, Status.OK.value,
     'проверка, "кавычки" и\nперенос строки', "1", "ok"),
]

PAGES = [
    (1, 1, "0001_p001.jpg", Status.OK.value, "", None),
    (1, 2, "0001_p002.jpg", Status.OK.value, "slow decode", None),
    (1, 3, "0001_p003.jpg", Status.FAILURE.value, "render failed", None),
    (2, 1, "", Status.FAILURE.value, "corrupt header", None),
    (3, 1, "0003_p001.jpg", Status.OK.value, "медленно, но «ок»", None),
    (3, 2, "0003_p002.jpg", Status.OK.value,
     "Extracted exactly at 00:01:33.37. Saved.", 93.37),
]

EXPECTED_CAPTURE_CELLS = ["", "", "", "", "", "00:01:33.37"]


def seed_database(db_path):
    controller = SQLiteDatabaseController(db_path)
    for file_id, rel_path, ext, pages, status, comment, range_str, range_status in FILES:
        controller.handle_file_started(file_id, rel_path, ext, "TestPipeline")
        controller.handle_file_completed(
            file_id,
            FileSummary(
                total_discovered_pages=pages,
                applied_range_string=range_str,
                range_status_code=range_status,
                final_aggregate_status=status,
                final_aggregate_comment=comment,
            ),
        )
    for parent_id, page_number, filename, status, comment, capture_seconds in PAGES:
        controller.handle_frame_saved(
            parent_id,
            PageResult(page_number, filename, status, comment,
                       capture_seconds=capture_seconds),
        )
    controller.close()


@pytest.fixture
def exported(tmp_path):
    db_path = tmp_path / "application_state.db"
    seed_database(db_path)

    exporter = SQLiteDataExporter(db_path)
    exporter.export_all_formats(tmp_path)
    yield tmp_path
    exporter.close()


def read_csv(path):
    with open(path, newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def find_one(directory, pattern):
    matches = list(directory.glob(pattern))
    assert len(matches) == 1, f"expected exactly one {pattern}, got {matches}"
    return matches[0]



LEAKY_MESSAGE = (
    "Catastrophic Image error: cannot identify image file "
    "'\\\\\\\\?\\\\C:\\\\pics\\\\holiday.jpeg'"
)


def test_no_exported_cell_can_carry_the_long_path_prefix(tmp_path):
    db_path = tmp_path / "leak_check.db"
    controller = SQLiteDatabaseController(db_path)
    controller.handle_file_started(1, "pics/holiday.jpeg", ".jpeg", "StaticImagePipeline")
    controller.handle_file_completed(
        1,
        FileSummary(
            total_discovered_pages=0,
            applied_range_string="",
            range_status_code=Status.FAILURE.value,
            final_aggregate_status=Status.FAILURE.value,
            final_aggregate_comment=LEAKY_MESSAGE,
        ),
    )
    controller.handle_frame_saved(1, PageResult(1, "", Status.FAILURE.value, LEAKY_MESSAGE))
    controller.close()

    exporter = SQLiteDataExporter(db_path)
    try:
        exporter.export_all_formats(tmp_path)
    finally:
        exporter.close()

    registry = read_csv(find_one(tmp_path, "file_registry_*.csv"))
    pages = read_csv(find_one(tmp_path, "page_log_*.csv"))
    assert "\\\\?\\" not in registry[0]["final_aggregate_comment"]
    assert "\\\\?\\" not in pages[0]["execution_comment"]
    assert "holiday.jpeg" in registry[0]["final_aggregate_comment"]

    workbook = openpyxl.load_workbook(find_one(tmp_path, "database_export_*.xlsx"))
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows(values_only=True):
            for cell in row:
                assert "\\\\?\\" not in str(cell)



def test_csv_registry_round_trips_seeded_content(exported):
    rows = read_csv(find_one(exported, "file_registry_*.csv"))

    assert [r["relative_file_path"] for r in rows] == [f[1] for f in FILES]
    for row, (file_id, _rel_path, ext, pages, status, comment, range_str,
              range_status) in zip(rows, FILES, strict=True):
        assert row["unique_file_id"] == str(file_id)
        assert row["original_extension"] == ext
        assert row["total_discovered_pages"] == str(pages)
        assert row["final_aggregate_status"] == status
        assert row["final_aggregate_comment"] == comment
        assert row["applied_range_string"] == range_str
        assert row["range_status_code"] == range_status
    assert set(rows[0]) == set(DatabaseFileRegistry.model_fields)


def test_csv_page_log_round_trips_seeded_content(exported):
    rows = read_csv(find_one(exported, "page_log_*.csv"))

    assert len(rows) == len(PAGES)
    for row, (parent_id, page_number, filename, status, comment, _), expected_hms \
            in zip(rows, PAGES, EXPECTED_CAPTURE_CELLS, strict=True):
        assert row["parent_file_id"] == str(parent_id)
        assert row["page_or_frame_number"] == str(page_number)
        assert row["saved_filename"] == filename
        assert row["execution_status"] == status
        assert row["execution_comment"] == comment
        assert row["capture_timestamp"] == expected_hms
        assert row["llm_answer_json"] == ""
    assert set(rows[0]) == set(DatabasePageLog.model_fields)



def test_xlsx_round_trips_seeded_content(exported):
    workbook = openpyxl.load_workbook(find_one(exported, "database_export_*.xlsx"))

    assert workbook.sheetnames == ["Master Registry", "Page Log - Part 1"]

    master = list(workbook["Master Registry"].iter_rows(values_only=True))
    assert list(master[0]) == list(DatabaseFileRegistry.model_fields)
    by_column = [dict(zip(master[0], row, strict=True)) for row in master[1:]]
    assert [r["relative_file_path"] for r in by_column] == [f[1] for f in FILES]
    assert [r["unique_file_id"] for r in by_column] == [f[0] for f in FILES]
    assert [r["total_discovered_pages"] for r in by_column] == [f[3] for f in FILES]
    assert [r["final_aggregate_status"] for r in by_column] == [f[4] for f in FILES]

    pages = list(workbook["Page Log - Part 1"].iter_rows(values_only=True))
    assert list(pages[0]) == list(DatabasePageLog.model_fields)
    page_rows = [dict(zip(pages[0], row, strict=True)) for row in pages[1:]]
    assert len(page_rows) == len(PAGES)
    assert [r["saved_filename"] for r in page_rows] == [p[2] or None for p in PAGES]
    assert [r["execution_status"] for r in page_rows] == [p[3] for p in PAGES]
    assert [r["capture_timestamp"] for r in page_rows] == \
        [cell or None for cell in EXPECTED_CAPTURE_CELLS]


def test_xlsx_page_log_splits_into_parts_at_row_limit(tmp_path, monkeypatch):
    monkeypatch.setattr(data_exporter, "EXCEL_MAX_ROWS", 5)

    db_path = tmp_path / "application_state.db"
    controller = SQLiteDatabaseController(db_path)
    controller.handle_file_started(1, "bulk/scan.pdf", ".pdf", "TestPipeline")
    for page_number in range(1, 13):
        controller.handle_frame_saved(
            1, PageResult(page_number, f"0001_p{page_number:03d}.jpg", Status.OK.value, "")
        )
    controller.close()

    exporter = SQLiteDataExporter(db_path)
    exporter.export_all_formats(tmp_path)
    exporter.close()

    workbook = openpyxl.load_workbook(find_one(tmp_path, "database_export_*.xlsx"))
    part_names = [n for n in workbook.sheetnames if n.startswith("Page Log")]
    assert part_names == ["Page Log - Part 1", "Page Log - Part 2", "Page Log - Part 3"]

    headers = list(DatabasePageLog.model_fields)
    collected = []
    for name in part_names:
        rows = list(workbook[name].iter_rows(values_only=True))
        assert list(rows[0]) == headers, f"{name} is missing its header row"
        collected.extend(dict(zip(headers, row, strict=True)) for row in rows[1:])

    assert [len(list(workbook[n].rows)) - 1 for n in part_names] == [5, 5, 2]
    assert [r["page_or_frame_number"] for r in collected] == list(range(1, 13))
