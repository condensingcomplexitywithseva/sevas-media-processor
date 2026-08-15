# Copyright 2026 Vsevolod Belonogov
# SPDX-License-Identifier: Apache-2.0


import csv
import sqlite3
import time
import logging
from datetime import datetime
from pathlib import Path
from collections.abc import Callable
from sqlmodel import Session, create_engine, select, col

from db_schema import DatabaseFileRegistry, DatabasePageLog
from fs_utils import get_safe_path
from range_parsers import truncate_visual_ranges

exporter_logger = logging.getLogger("DataExporter")

EXCEL_MAX_ROWS = 1_000_000


class SQLiteDataExporter:

    def __init__(self, target_database_path: Path, max_retries: int = 4):
        self.database_path = target_database_path
        self.max_retries = max_retries

        safe_database_path = get_safe_path(self.database_path)
        self.sql_engine = create_engine(
            "sqlite://", creator=lambda: sqlite3.connect(safe_database_path), echo=False
        )

    def close(self) -> None:
        try:
            self.sql_engine.dispose()
        except Exception:
            exporter_logger.warning("Could not cleanly dispose the export database engine.")

    def _execute_with_lock_protection(self, export_function: Callable[[], None],
                                      target_file_name: str, output_dir: Path) -> None:
        panic_path = output_dir / f"_LOCKED_ERROR_{target_file_name}.txt"

        for attempt in range(self.max_retries):
            try:
                export_function()

                try:
                    safe_panic = Path(get_safe_path(panic_path))
                    if safe_panic.exists():
                        safe_panic.unlink()
                except OSError:
                    pass
                return

            except (PermissionError, OSError) as os_error:
                if attempt < self.max_retries - 1:
                    exporter_logger.warning(
                        f"Target file '{target_file_name}' is locked by the OS. "
                        f"Attempt {attempt + 1}/{self.max_retries}."
                    )

                    try:
                        with open(get_safe_path(panic_path), "w", encoding="utf-8") as panic_file:
                            panic_file.write(f"Cannot overwrite {target_file_name}. Is it open in another program?")
                    except OSError:
                        pass

                    for remaining_seconds in range(30, 0, -10):
                        exporter_logger.warning(f"File locked. Retrying export in {remaining_seconds} seconds...")
                        time.sleep(10)

                else:
                    fatal_msg = (f"CRITICAL: Exhausted all retries. "
                                 f"Cannot access {target_file_name}. Details: {os_error}")
                    exporter_logger.critical(fatal_msg)
                    raise PermissionError(fatal_msg) from os_error

    def export_all_formats(self, output_directory_path: Path) -> None:
        Path(get_safe_path(output_directory_path)).mkdir(parents=True, exist_ok=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        exporter_logger.info(f"Initiating SQLite data export sequence (Snapshot: {timestamp})...")

        self._export_to_csv(output_directory_path, timestamp)
        self._export_to_xlsx(output_directory_path, timestamp)

    def export_csv(self, output_directory_path: Path, timestamp: str | None = None) -> None:
        Path(get_safe_path(output_directory_path)).mkdir(parents=True, exist_ok=True)
        timestamp = timestamp or datetime.now().strftime("%Y%m%d_%H%M%S")
        self._export_to_csv(output_directory_path, timestamp)

    def _export_to_csv(self, output_directory_path: Path, timestamp: str) -> None:
        master_registry_path = output_directory_path / f"file_registry_{timestamp}.csv"
        page_log_path = output_directory_path / f"page_log_{timestamp}.csv"

        def write_master_registry() -> None:
            with open(get_safe_path(master_registry_path), "w", newline="", encoding="utf-8") as csv_file:
                headers = list(DatabaseFileRegistry.model_fields.keys())
                writer = csv.DictWriter(csv_file, fieldnames=headers, quoting=csv.QUOTE_ALL)
                writer.writeheader()

                with Session(self.sql_engine) as session:
                    statement = select(DatabaseFileRegistry).order_by(col(DatabaseFileRegistry.unique_file_id))
                    results = session.exec(statement).all()
                    for row in results:
                        data = row.model_dump()
                        if data.get("applied_range_string"):
                            parts = [p.strip() for p in data["applied_range_string"].split(",")]
                            data["applied_range_string"] = truncate_visual_ranges(parts)
                        writer.writerow(data)
            exporter_logger.info(f"Successfully exported master registry to: {master_registry_path.name}")

        def write_page_log() -> None:
            with open(get_safe_path(page_log_path), "w", newline="", encoding="utf-8") as csv_file:
                headers = list(DatabasePageLog.model_fields.keys())
                writer = csv.DictWriter(csv_file, fieldnames=headers, quoting=csv.QUOTE_ALL)
                writer.writeheader()

                with Session(self.sql_engine) as session:
                    statement = select(DatabasePageLog).order_by(col(DatabasePageLog.primary_database_id))
                    results = session.exec(statement).yield_per(1000)
                    for row in results:
                        writer.writerow(row.model_dump())
            exporter_logger.info(f"Successfully exported page log to: {page_log_path.name}")

        self._execute_with_lock_protection(write_master_registry, "file_registry.csv", output_directory_path)
        self._execute_with_lock_protection(write_page_log, "page_log.csv", output_directory_path)

    def _export_to_xlsx(self, output_directory_path: Path, timestamp: str) -> None:
        try:
            import openpyxl
            from openpyxl.worksheet.worksheet import Worksheet
        except ImportError:
            exporter_logger.warning("openpyxl is not installed. Skipping XLSX export.")
            return

        excel_path = output_directory_path / f"database_export_{timestamp}.xlsx"

        def write_excel_workbook() -> None:
            workbook = openpyxl.Workbook()

            active_sheet = workbook.active
            if isinstance(active_sheet, Worksheet):
                master_sheet = active_sheet
            else:
                master_sheet = workbook.create_sheet()

            master_sheet.title = "Master Registry"
            master_headers = list(DatabaseFileRegistry.model_fields.keys())
            master_sheet.append(master_headers)

            page_headers = list(DatabasePageLog.model_fields.keys())
            sheet_index = 1

            current_page_sheet = workbook.create_sheet(title=f"Page Log - Part {sheet_index}")

            if not isinstance(current_page_sheet, Worksheet):
                raise TypeError("Failed to generate a valid Excel Worksheet object.")

            current_page_sheet.append(page_headers)

            with Session(self.sql_engine) as session:
                master_statement = select(DatabaseFileRegistry).order_by(col(DatabaseFileRegistry.unique_file_id))
                master_results = session.exec(master_statement).all()
                for row in master_results:
                    data = row.model_dump()
                    if data.get("applied_range_string"):
                        parts = [p.strip() for p in data["applied_range_string"].split(",")]
                        data["applied_range_string"] = truncate_visual_ranges(parts)
                    master_sheet.append([data.get(header) for header in master_headers])

                page_statement = select(DatabasePageLog).order_by(col(DatabasePageLog.primary_database_id))
                page_results = session.exec(page_statement).yield_per(1000)

                row_counter = 0

                for row in page_results:
                    if row_counter >= EXCEL_MAX_ROWS:
                        sheet_index += 1
                        current_page_sheet = workbook.create_sheet(title=f"Page Log - Part {sheet_index}")

                        if not isinstance(current_page_sheet, Worksheet):
                            raise TypeError("Failed to generate a valid Excel Worksheet object during pagination.")

                        current_page_sheet.append(page_headers)
                        row_counter = 0

                    data = row.model_dump()
                    current_page_sheet.append([data.get(header) for header in page_headers])
                    row_counter += 1

            workbook.save(get_safe_path(excel_path))
            exporter_logger.info(f"Successfully exported database to Excel: {excel_path.name}")

        self._execute_with_lock_protection(write_excel_workbook, "database_export.xlsx", output_directory_path)
